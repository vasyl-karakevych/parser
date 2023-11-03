import requests
from bs4 import BeautifulSoup
import openpyxl
from product import Product
import os
from sheet import Sheet

def Open_url(list_products, url):
    # Відправка запиту до сторінки і отримання її вмісту
    response = requests.get(url)

    # Перевірка статус-коду відповіді
    if response.status_code == 200:
        # Парсимо HTML-сторінку
        soup = BeautifulSoup(response.text, 'html.parser')

        # Знаходимо всі елементи з назвою товару і ціною
        product_names = soup.find_all('h2', class_='font-headline text-lg font-bold leading-6 line-clamp-3 md:text-xl md:leading-8')
        prices = soup.find_all('div', class_='text-3xl font-bold leading-8')
        p_available = soup.find_all('div', class_='leading-tight')
        # font-semibold

        # Виводимо дані на екран
        for row, (name, price, p_available) in enumerate(zip(product_names, prices, p_available), start=1):
            product_name = name.text.strip()
            product_price = price.text.strip()
            product_available = p_available.text.strip()

            if product_available == "": continue
            # print(product_available)
            product_price = ''.join(product_price[0:-3].split())

        # Видаляємо якщо пише у назві OUTLET   
            if product_name[-7:-1] == "Outlet": continue

            products = Product(product_name, product_price)
            list_products.append(products)

        
    else:
        print(f'Помилка отримання сторінки. Статус-код: {response.status_code}')
# URL сторінки, яку потрібно спарсити

def Fiter_List(list_products):
    # Створюємо словник для відстеження унікальних імен
    unique_names = {}

    # Фільтруємо список, залишаючи тільки об'єкти з унікальними іменами
    filtered_products = []

    for product in list_products:
        if product.name not in unique_names:
            unique_names[product.name] = True
            filtered_products.append(product)
    list_products = filtered_products
    return list_products

def Write_Exel(list_products, type):
    # RemoveExcelFile("komputronik.wlsx")

    # workbook = openpyxl.Workbook()
    # sheet = workbook.active
    # Створити нову вкладку (лист) і назвати її
    # new_sheet = workbook.create_sheet(type)
    workbook = openpyxl.load_workbook('komputronik.xlsx')
    sheet = workbook[type]

    # Встановити ширину стовпця (наприклад, стовпця A) в пікселях
    sheet.column_dimensions['B'].width = 70  

    # Очистіть всі дані на аркуші
    for row in sheet:
        for cell in row:
            cell.value = None

    for row, (product) in enumerate(list_products, start=1):
        sheet.cell(row=row, column=1, value=row)
        sheet.cell(row=row, column=2, value=product.getName())
        sheet.cell(row=row, column=3, value='=')
        # sheet.cell(row=row, column=2, value=product.getPrice())
        if type == 'Laptops': 
            sheet.cell(row=row, column=4, value=product.getPriceWithDelivery() + 400)
        else:
            sheet.cell(row=row, column=4, value=product.getPriceWithDelivery())

        sheet.cell(row=row, column=5, value='грн.')
        
    workbook.save('komputronik.xlsx')

def RemoveExcelFile(file_name):
    file_name = "komputronik.xlsx"  

    if os.path.exists(file_name):
        os.remove(file_name)
        print(f"Файл {file_name} було видалено.")
    else:
        print(f"Файл {file_name} не існує.")
   
def parcing_url(url_, type, count):
    list_products = []
    for i in range(1,count):
        url = f'{url_}{i}'
        print(f'Parcing {url}')
        Open_url(list_products, url)
    list_products = Fiter_List(list_products)
    list_products = sorted(list_products, key=lambda x: x.price)
    Write_Exel(list_products, type)
    print(f'Scan and wrote= {len(list_products)} elements')


print('parcing Grafik cards')
parcing_url('https://www.komputronik.pl/category/1099/karty-graficzne.html?showBuyActiveOnly=0&p=', 'Karty', 40)

print('Parcing laptops')
parcing_url('https://www.komputronik.pl/category/5022/laptopy.html?showBuyActiveOnly=0&p=', 'Laptops', 70)



# parcing_url(url_list)
# print(list_products)

